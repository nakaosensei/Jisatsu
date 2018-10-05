import openpyxl

dado = openpyxl.load_workbook('ListaMacrofitas.xlsx')
planilha = dado.sheetnames[0]
sheet = dado[planilha]
pesquisa = []

for row in sheet.iter_rows('A1:A1048576'):
	for cell in row:
		if cell.value is None:
			break
		nome = cell.value
		nome = nome.split(" ")
		pesquisa.append(nome[0]+"_"+nome[1])

print(pesquisa)